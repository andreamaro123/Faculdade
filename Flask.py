"""
Modulo de escaneamento de rede usando Flask e Nmap.

Este programa usa Flask para criar uma interface web que permite escanear redes 
usando Nmap e salvar os resultados em um arquivo Excel formatado.
"""

from flask import Flask, render_template, request, send_from_directory, jsonify
import nmap
import os
import pandas as pd
import threading
from time import sleep
from flask_socketio import SocketIO, emit
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

app = Flask(__name__)
socketio = SocketIO(app)

SCAN_RESULTS_DIR = "scan_results"
os.makedirs(SCAN_RESULTS_DIR, exist_ok=True)
# Funcao para escanear a rede usando Nmap
def scan_network(ip_range, progress_callback):
    nm = nmap.PortScanner()
    active_hosts = []
    
    try:
        nm.scan(hosts=ip_range, arguments='-sS -T5 -O -p 80,443,22,21,53,110,143,3389,8080')
    except Exception as e:
        print(f"Erro ao escanear: {e}")
        return None

    hosts = nm.all_hosts()
    for idx, host in enumerate(hosts):
        progress = int((idx + 1) / len(hosts) * 100)
        progress_callback(progress)
        sleep(0.2)

        if nm[host].state() == "up":
            # Verifica se o sistema operacional foi identificado pelo Nmap
            if 'osmatch' in nm[host]:
                os_match = nm[host]['osmatch']
                os_info = os_match[0] if os_match else 'Desconhecido'
            else:
                os_info = 'Desconhecido'

            host_info = {
                'IP': host,
                'Hostname': nm[host].hostname(),
                'Status': 'Ativo',
                'OS': os_info,  # Usando a variavel que contem a informacao do SO
                'Ports': []
            }

            for proto in nm[host].all_protocols():
                lport = nm[host][proto].keys()
                for port in lport:
                    host_info['Ports'].append(f"{port}/{proto}")

            active_hosts.append(host_info)

            
            socketio.emit('new_ip', {'ip': host})

    # Gerar o arquivo Excel
    df = pd.DataFrame(active_hosts)
    file_path = os.path.join(SCAN_RESULTS_DIR, "inventario_atual.xlsx")
    df.to_excel(file_path, index=False)
    
    # Formata o arquivo Excel
    wb = load_workbook(file_path)
    ws = wb.active
    headers = ['IP', 'Hostname', 'Status', 'Sistema Operacional', 'Ports']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(file_path)

    return file_path

# Funcao que atualiza o progresso
def update_progress(value):
    socketio.emit('progress', {'progress': value})

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/scan", methods=["POST"])
def scan():
    try:
        data = request.get_json()   # Espera receber os dados como JSON
        ip_range = data.get("ip_range")

        if not ip_range:
            return jsonify({"error": "Faixa de IP não fornecida."}), 400

        # Funcao para iniciar o scan
        def start_scan():
            file_path = scan_network(ip_range, update_progress)
            if file_path:
                download_url = f"/download/{os.path.basename(file_path)}"
                socketio.emit('scan_complete', {'file_path': download_url})
            else:
                socketio.emit('scan_complete', {'file_path': 'Erro no scan. Tente novamente.'})

        scan_thread = threading.Thread(target=start_scan)
        scan_thread.start()

        return jsonify({"message": "Scan iniciado."})  # Responde com mensagem de sucesso

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/download/<filename>")
def download(filename):
    """
    Permite o download do arquivo de resultados de escaneamento.
    """
    return send_from_directory(SCAN_RESULTS_DIR, filename)

if __name__ == "__main__":
    """
    Inicia o servidor Flask com suporte ao WebSocket via SocketIO.
    """
    socketio.run(app, debug=True)
